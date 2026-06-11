# -------------------------------------------------------------------------------
# Name:        export/excel_pandas.py
# Purpose:     Export triplet DataFrames to Excel format
# -------------------------------------------------------------------------------
import os
import logging

from io import BytesIO

import pandas

from triplets.tools import triplets_to_tableviews

logger = logging.getLogger(__name__)


def export_to_excel(data, path=None, multivalue=True, export_to_memory=False, single_file=False, filename=None, apply_formatting=True):
    """Export triplet data to Excel file(s), with each type on a separate sheet.

    Parameters
    ----------
    data : pandas.DataFrame
        Triplet dataset containing RDF data.
    path : str, optional
        Directory path to save Excel file(s), or file path when single_file=True.
    multivalue : bool, default True
        If True, aggregate duplicate (ID, KEY) pairs into lists.
    export_to_memory : bool, default False
        If True, return BytesIO objects; if False, save to disk.
    single_file : bool, default False
        If True, export all data to a single file instead of one file per INSTANCE_ID.
    filename : str, optional
        Filename to use when single_file=True. If None, uses 'export.xlsx'.
    apply_formatting : bool, default True
        If True, apply column width and freeze panes formatting.

    Returns
    -------
    BytesIO, str, or list
        Depends on single_file and export_to_memory flags.
    """
    if single_file:
        if filename is None:
            filename = 'export.xlsx'

        tableviews = triplets_to_tableviews(data, multivalue=multivalue)
        output = BytesIO()
        output.name = filename

        with pandas.ExcelWriter(output, engine='openpyxl') as writer:
            for class_type, class_data in tableviews.items():
                class_data.to_excel(writer, sheet_name=class_type)
                if apply_formatting:
                    from openpyxl.utils import get_column_letter
                    sheet = writer.sheets[class_type]
                    for i in range(1, len(class_data.columns) + 2):
                        sheet.column_dimensions[get_column_letter(i)].width = 38
                    sheet.freeze_panes = 'B2'

        output.seek(0)

        if export_to_memory:
            return output
        else:
            if path is None:
                path = os.getcwd()
            if os.path.isdir(path) or not path.endswith('.xlsx'):
                export_path = os.path.join(path, filename)
            else:
                export_path = path
            with open(export_path, 'wb') as f:
                f.write(output.read())
            logger.info(f'Saved {export_path}')
            return filename
    else:
        labels = data.query("KEY == 'label'")
        exported_files = []

        for _, label in labels.iterrows():
            instance_data = data[data.INSTANCE_ID == label.INSTANCE_ID]
            tableviews = triplets_to_tableviews(instance_data, multivalue=multivalue)
            file_name = '{}.xlsx'.format(label.VALUE.split(".")[0])
            output = BytesIO()
            output.name = file_name

            with pandas.ExcelWriter(output, engine='openpyxl') as writer:
                for class_type, class_data in tableviews.items():
                    class_data.to_excel(writer, sheet_name=class_type)
                    if apply_formatting:
                        from openpyxl.utils import get_column_letter
                        sheet = writer.sheets[class_type]
                        for i in range(1, len(class_data.columns) + 2):
                            sheet.column_dimensions[get_column_letter(i)].width = 38
                        sheet.freeze_panes = 'B2'

            output.seek(0)
            exported_files.append(output)

        if export_to_memory:
            return exported_files
        else:
            if path is None:
                path = os.getcwd()
            exported_file_names = []
            for file_object in exported_files:
                export_path = os.path.join(path, file_object.name)
                with open(export_path, 'wb') as f:
                    file_object.seek(0)
                    f.write(file_object.read())
                exported_file_names.append(file_object.name)
                logger.info(f'Saved {export_path}')
            return exported_file_names
